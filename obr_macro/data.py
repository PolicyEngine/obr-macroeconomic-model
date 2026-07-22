"""Load OBR EFO data from Excel files."""

from pathlib import Path
import urllib.request
import json

import numpy as np
import pandas as pd
import openpyxl

# OBR March 2026 EFO download URLs
OBR_URLS = {
    "economy": "https://obr.uk/download/march-2026-economic-and-fiscal-outlook-detailed-forecast-tables-economy/",
    "receipts": "https://obr.uk/download/march-2026-economic-and-fiscal-outlook-detailed-forecast-tables-receipts/",
    "expenditure": "https://obr.uk/download/march-2026-economic-and-fiscal-outlook-detailed-forecast-tables-expenditure/",
    "aggregates": "https://obr.uk/download/march-2026-economic-and-fiscal-outlook-detailed-forecast-tables-aggregates/",
}


def _resolve_data_dir() -> Path:
    """Locate the data directory.

    In a git checkout, `data/` at the repo root is the historical location
    (gitignored; populated by the ensure_* downloads) — prefer it so existing
    workflows, tests, and the ons_cache keep working unchanged. In a pip
    install there is no repo root, so fall back to `obr_macro/_data/`, which
    ships the model code, model variables, and EFO workbooks as package data.
    """
    repo_data = Path(__file__).resolve().parent.parent / "data"
    if repo_data.is_dir():
        return repo_data
    return Path(__file__).resolve().parent / "_data"


DATA_DIR = _resolve_data_dir()

# OBR macroeconomic model code/variables download URLs (15 October 2025 version)
OBR_MODEL_URLS = {
    "model_code": "https://obr.uk/download/obr-macroeconomic-model-code/",
    "model_variables": "https://obr.uk/download/obr-macroeconomic-model-variables/",
}

# The OBR site sits behind Cloudflare; a bare "Mozilla/5.0" User-Agent gets
# 403-blocked, so a realistic full browser header set is required.
_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language": "en-GB,en;q=0.9",
    "Referer": "https://obr.uk/",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
}


def _download_validated(url: str, path: Path, kind: str) -> None:
    """Download `url` to `path` atomically, validating the payload first.

    The OBR site sits behind Cloudflare; a challenge/error page returns HTTP 200
    with HTML, and a truncated read yields a partial file. Either would be cached
    permanently as the "model file". Download to a temp path, check the status
    and content signature, then atomically rename into place.

    kind: "xlsx" (must start with the ZIP magic PK\\x03\\x04) or
          "text" (must be >10KB, contain '=', and not look like HTML).
    """
    req = urllib.request.Request(url, headers=_BROWSER_HEADERS)
    with urllib.request.urlopen(req, timeout=60) as resp:
        if getattr(resp, "status", 200) != 200:
            raise RuntimeError(
                f"{kind} download for {path.name} returned HTTP {resp.status}"
            )
        blob = resp.read()

    if kind == "xlsx":
        if not blob.startswith(b"PK\x03\x04"):
            raise RuntimeError(
                f"{path.name}: expected an xlsx (ZIP) payload but got "
                f"{blob[:16]!r} — likely a Cloudflare challenge page, not the file."
            )
    else:  # text
        head = blob.lstrip()[:64].lower()
        if len(blob) < 10_000 or b"=" not in blob or head.startswith(b"<"):
            raise RuntimeError(
                f"{path.name}: payload does not look like the model code "
                f"({len(blob)} bytes, starts {blob[:16]!r}) — likely a blocked/partial download."
            )

    tmp = path.with_suffix(path.suffix + ".part")
    tmp.write_bytes(blob)
    tmp.replace(path)  # atomic on the same filesystem


def ensure_model_code() -> Path:
    """Download the OBR model code txt and variables xlsx if not present.

    Returns the Path to the model code txt file. Both files are saved into
    DATA_DIR. Existing files are not re-downloaded.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    files = {
        "model_code": DATA_DIR / "obr_model_code_october_2025.txt",
        "model_variables": DATA_DIR / "obr_model_variables_october_2025.xlsx",
    }
    kinds = {"model_code": "text", "model_variables": "xlsx"}

    for key, path in files.items():
        if not path.exists():
            print(f"Downloading {key}...")
            _download_validated(OBR_MODEL_URLS[key], path, kinds[key])

    return files["model_code"]


def ensure_downloaded() -> dict[str, Path]:
    """Download OBR EFO Excel files if not present."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    files = {
        "economy": DATA_DIR / "obr_efo_march_2026_economy.xlsx",
        "receipts": DATA_DIR / "obr_efo_march_2026_receipts.xlsx",
        "expenditure": DATA_DIR / "obr_efo_march_2026_expenditure.xlsx",
        "aggregates": DATA_DIR / "obr_efo_march_2026_aggregates.xlsx",
    }

    for key, path in files.items():
        if not path.exists():
            print(f"Downloading {key}...")
            _download_validated(OBR_URLS[key], path, "xlsx")

    return files


def load_variable_definitions() -> list[dict]:
    """Load the OBR variable definitions from JSON."""
    var_file = DATA_DIR / "obr_model_variables.json"
    if var_file.exists():
        with open(var_file) as f:
            return json.load(f)
    return []


def _assert_headers(
    wb: openpyxl.Workbook, sheet: str, checks: dict[tuple[int, int], str]
) -> None:
    """Sanity-check header text on a sheet read by absolute column position.

    ``checks`` maps (row, col) -> expected substring (case-insensitive). Raises
    a clear error if the sheet layout changed, instead of silently reading the
    wrong columns into the wrong variables.
    """
    ws = wb[sheet]
    for (r, c), expected in checks.items():
        actual = str(ws.cell(r, c).value or "").strip()
        if expected.lower() not in actual.lower():
            raise RuntimeError(
                f"OBR EFO sheet {sheet!r} layout changed: expected header cell "
                f"({r},{c}) to contain {expected!r} but found {actual!r}. "
                "The absolute column mapping in load_obr_data needs updating."
            )


def _read_quarterly_table(
    wb: openpyxl.Workbook,
    sheet: str,
    period_col: int,
    data_cols: dict[str, int],
    min_row: int = 4,
    max_row: int = 200,
) -> dict[str, pd.Series]:
    """Read quarterly data from an OBR Excel sheet."""
    ws = wb[sheet]
    result = {}

    for var_name, col in data_cols.items():
        data = {}
        for r in range(min_row, max_row + 1):
            period = ws.cell(r, period_col).value
            if period is None:
                continue
            period = str(period).strip()
            if len(period) != 6 or "Q" not in period:
                continue
            val = ws.cell(r, col).value
            if isinstance(val, (int, float)) and np.isfinite(val):
                data[pd.Period(period, freq="Q")] = float(val)

        if data:
            result[var_name] = pd.Series(data).sort_index()

    return result


def load_obr_data(merge_snapshot: bool = True) -> pd.DataFrame:
    """Load all OBR forecast data into a DataFrame.

    Returns a DataFrame with quarterly PeriodIndex and columns for each variable.
    Values are in model units (£m for monetary values, % for rates, index for prices).
    """
    files = ensure_downloaded()

    all_data = {}

    # Economy tables
    wb = openpyxl.load_workbook(str(files["economy"]), data_only=True)

    # Guard the absolute-column reads below: assert the expected table titles
    # and key column headers so a re-laid-out EFO workbook fails loudly rather
    # than loading the wrong columns into the wrong variables.
    _assert_headers(
        wb,
        "1.1",
        {(2, 2): "1.1 GDP expend", (3, 3): "Private consum", (3, 4): "Government con"},
    )
    _assert_headers(wb, "1.2", {(2, 2): "1.2 GDP expend", (3, 3): "Private consum"})
    _assert_headers(wb, "1.3", {(2, 2): "1.3 GDP income", (3, 3): "Labour income"})
    _assert_headers(
        wb,
        "1.6",
        {(2, 2): "1.6 Labour mar", (3, 3): "Employment", (3, 6): "ILO unemployment"},
    )
    _assert_headers(wb, "1.7", {(2, 2): "1.7 Inflation", (3, 3): "Year-on-year"})
    _assert_headers(
        wb, "1.9", {(2, 2): "1.9 Market", (3, 3): "Bank Rate", (3, 4): "Long-term"}
    )
    _assert_headers(wb, "1.11", {(2, 2): "1.11 Household"})
    _assert_headers(wb, "1.12", {(2, 2): "1.12 Household", (3, 3): "Labour Income"})
    _assert_headers(wb, "1.16", {(2, 2): "1.16 Housing", (3, 3): "House price in"})
    _assert_headers(wb, "1.8", {(2, 2): "1.8 Balance of", (3, 3): "Trade balance"})
    _assert_headers(
        wb, "1.10", {(2, 2): "1.10 Financial", (3, 3): "% GDP", (3, 8): "£ billion"}
    )
    _assert_headers(wb, "1.14", {(2, 2): "1.14 OBR centr"})

    # Table 1.1: Real GDP expenditure (£bn)
    data = _read_quarterly_table(
        wb,
        "1.1",
        period_col=2,
        data_cols={
            "CONS": 3,  # Private consumption
            "CGG": 4,  # Government consumption
            "IF": 5,  # Fixed investment
            "IBUS": 6,  # Business investment
            "IH": 7,  # Dwellings investment
            "GGI": 8,  # Government investment
            "PCIH": 9,  # Public corps dwellings
            "VAL": 10,  # Net acquisition of valuables
            "DINV": 12,  # Change in inventories
            "X": 14,  # Exports
            "M": 16,  # Imports
            "SDE": 17,  # Statistical discrepancy
            "GDPM": 18,  # Real GDP
            "NOGVA": 19,  # Non-oil GVA
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000  # £bn → £m

    # Table 1.2: Nominal GDP expenditure (£bn)
    data = _read_quarterly_table(
        wb,
        "1.2",
        period_col=2,
        data_cols={
            "CONSPS": 3,  # Nominal consumption
            "CGGPS": 4,  # Nominal government consumption
            "IFPS": 5,  # Nominal investment
            "GGIPS": 6,  # Nominal government investment
            "VALPS": 7,  # Nominal valuables
            "DINVPS": 9,  # Nominal change in inventories
            "XPS": 11,  # Nominal exports
            "MPS": 13,  # Nominal imports
            "GDPMPS": 15,  # Nominal GDP
            "GNI": 16,  # Gross national income
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.3: GDP income (£bn)
    data = _read_quarterly_table(
        wb,
        "1.3",
        period_col=2,
        data_cols={
            "FYEMP": 3,  # Labour income (compensation)
            "FYCPR": 4,  # Non-oil PNFC profits
            "OTHFYINC": 5,  # Other income
            "GVAFC": 6,  # GVA at factor cost
            "BPAPS": 7,  # Taxes less subsidies on products
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.6: Labour market
    data = _read_quarterly_table(
        wb,
        "1.6",
        period_col=2,
        data_cols={
            "ETLFS": 3,  # Employment (millions)
            "ETRATE": 4,  # Employment rate (%)
            "EEES": 5,  # Employees (millions)
            "ULFSU": 6,  # Unemployment (millions)
            "LFSUR": 7,  # Unemployment rate (%)
            "PART16": 8,  # Participation rate (%)
            "AVH": 9,  # Average hours
            "HWA": 10,  # Total hours (millions)
            "LABSH": 11,  # Labour income share (%)
            "COMPSH": 12,  # Compensation share (%)
            "COMP": 13,  # Compensation of employees (£bn)
            "WFP": 14,  # Wages and salaries (£bn)
            "EMPSC": 15,  # Employer social contributions (£bn)
            "MIXINC": 16,  # Mixed income (£bn)
            "AWEGR": 17,  # AWE growth (%)
            "AWEI": 18,  # AWE index
            "AHEI": 19,  # Avg hourly earnings index
            "PRODH": 20,  # Productivity per hour index
            "PRODW": 21,  # Productivity per worker index
            "RPW": 22,  # Real product wage
            "RCW": 23,  # Real consumption wage
            "RWEI": 24,  # Real weekly earnings index
        },
    )
    all_data["ETLFS"] = data["ETLFS"] * 1000  # millions → thousands
    all_data["EEES"] = data["EEES"] * 1000
    all_data["ULFSU"] = data["ULFSU"] * 1000
    all_data["LFSUR"] = data["LFSUR"]
    all_data["ETRATE"] = data["ETRATE"]
    all_data["PART16"] = data["PART16"]
    all_data["AVH"] = data["AVH"]
    all_data["HWA"] = data["HWA"]
    all_data["LABSH"] = data["LABSH"]
    all_data["COMPSH"] = data["COMPSH"]
    all_data["COMP"] = data["COMP"] * 1000
    all_data["WFP"] = data["WFP"] * 1000
    all_data["EMPSC"] = data["EMPSC"] * 1000
    all_data["MIXINC"] = data["MIXINC"] * 1000
    all_data["AWEGR"] = data["AWEGR"]
    all_data["AWEI"] = data["AWEI"]
    all_data["AHEI"] = data["AHEI"]
    all_data["PRODH"] = data["PRODH"]
    all_data["PRODW"] = data["PRODW"]
    all_data["RPW"] = data["RPW"]
    all_data["RCW"] = data["RCW"]
    all_data["RWEI"] = data["RWEI"]

    # Table 1.7: Prices - growth rates
    data = _read_quarterly_table(
        wb,
        "1.7",
        period_col=2,
        data_cols={
            "RPIGR": 3,  # RPI growth
            "RPIXGR": 4,  # RPIX growth
            "CPIGR": 5,  # CPI growth
            "CPIHGR": 6,  # CPIH growth
            "OOHGR": 7,  # OOH growth
            "MIPGR": 8,  # Mortgage interest growth
            "RENTGR": 9,  # Actual rents growth
            "PRVRENTGR": 10,  # Private rentals growth
            "PCEGR": 11,  # Consumer deflator growth
            "PGDPGR": 12,  # GDP deflator growth
        },
    )
    all_data.update(data)

    # Table 1.7: Prices - indices
    data = _read_quarterly_table(
        wb,
        "1.7",
        period_col=2,
        data_cols={
            "RPI": 13,  # RPI index
            "RPIX": 14,  # RPIX index
            "CPI": 15,  # CPI index (2015=100)
            "CPIH": 16,  # CPIH index
            "OOH": 17,  # OOH index
            "MIP": 18,  # Mortgage interest index
            "RENT": 19,  # Actual rents index
            "PRVRENT": 20,  # Private rentals index
            "PCE": 21,  # Consumer expenditure deflator
            "PGDP": 22,  # GDP deflator
        },
    )
    all_data.update(data)

    # Table 1.9: Market assumptions
    data = _read_quarterly_table(
        wb,
        "1.9",
        period_col=2,
        data_cols={
            "R": 3,  # Bank Rate (%)
            "RL": 4,  # Gilt yield (%)
            "RMORT": 5,  # Mortgage rate (%)
            "RDEP": 6,  # Deposit rate (%)
            "RX": 7,  # Sterling ERI
            "RXD": 8,  # USD/GBP
            "RXE": 9,  # EUR/GBP
            "PBRENT": 10,  # Oil price ($/bbl)
            "PGAS": 11,  # Gas price (£)
            "PEQTY": 12,  # Equity prices
        },
    )
    all_data.update(data)

    # Table 1.11: Balance sheets (£bn)
    data = _read_quarterly_table(
        wb,
        "1.11",
        period_col=2,
        data_cols={
            "HHPHYSA": 3,  # HH physical assets
            "HHFINA": 4,  # HH financial assets
            "HHLIAB": 5,  # HH liabilities
            "HHSECLIAB": 6,  # HH secured liabilities
            "HHOTHLIAB": 7,  # HH other liabilities
            "HHNW": 8,  # HH total net worth
            "HHDIY": 9,  # HH disposable income (for ratios)
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.12: Household income (£bn)
    data = _read_quarterly_table(
        wb,
        "1.12",
        period_col=2,
        data_cols={
            "LABINC": 3,  # Labour income
            "EMPCOMP": 4,  # Employee compensation
            "MIXINC12": 5,  # Mixed income
            "EMPSC12": 6,  # Employer social contributions
            "NONLABINC": 7,  # Non-labour income
            "NETBENFIT": 8,  # Net benefits and taxes
            "HHDI": 9,  # Household disposable income
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.16: Housing
    data = _read_quarterly_table(
        wb,
        "1.16",
        period_col=2,
        data_cols={
            "APH": 3,  # House price index
            "APHGR": 4,  # House price growth
            "PROPTX": 5,  # Property transactions (000s)
            "PEHS": 6,  # Private housing starts
            "PEHC": 7,  # Private housing completions
            "HSTOCK": 8,  # Housing stock (000s)
            "HNETADD": 9,  # Net housing additions
            "HTURN": 10,  # Turnover rate
        },
    )
    all_data.update(data)

    # Table 1.8: Balance of payments (£bn)
    data = _read_quarterly_table(
        wb,
        "1.8",
        period_col=2,
        data_cols={
            "TB": 3,  # Trade balance
            "NIPD": 5,  # Investment income balance
            "CB": 8,  # Current account balance
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.10: Financial balances (% GDP)
    data = _read_quarterly_table(
        wb,
        "1.10",
        period_col=2,
        data_cols={
            "NAFHH_PCT": 3,  # Household net lending (% GDP)
            "NAFCO_PCT": 4,  # Corporate net lending (% GDP)
            "NAFGG_PCT": 5,  # Government net lending (% GDP)
            "NAFROW_PCT": 6,  # ROW net lending (% GDP)
        },
    )
    all_data.update(data)

    # Table 1.10: Financial balances (£bn)
    data = _read_quarterly_table(
        wb,
        "1.10",
        period_col=2,
        data_cols={
            "NAFHH": 8,  # Household net lending
            "NAFCO": 9,  # Corporate net lending
            "NAFGG": 10,  # Government net lending
            "NAFROW": 11,  # ROW net lending
        },
    )
    for k, v in data.items():
        all_data[k] = v * 1000

    # Table 1.14: Output gap
    data = _read_quarterly_table(
        wb,
        "1.14",
        period_col=2,
        data_cols={
            "GAP": 3,  # Output gap (%)
        },
    )
    all_data.update(data)

    # Build DataFrame efficiently (avoid fragmentation)
    all_periods = set()
    for series in all_data.values():
        all_periods.update(series.index)

    index = pd.PeriodIndex(sorted(all_periods), freq="Q")

    # Create all columns at once to avoid fragmentation
    df = pd.DataFrame(all_data, index=index)

    # Derive additional variables from identities
    df = _derive_variables(df)

    # Merge the vendored ONS exogenous-input snapshot (Stage 1c): the
    # disaggregated National-Accounts / fiscal series the equations need but the
    # EFO tables do not publish. Held flat beyond the data for forecast periods.
    # Skipped via merge_snapshot=False when deriving what still needs pulling.
    if merge_snapshot:
        df = _merge_ons_snapshot(df)

    return df


def _merge_ons_snapshot(df: pd.DataFrame) -> pd.DataFrame:
    """Add the ONS exogenous-input series (obr_macro/seeds/...) to the frame."""
    snap_path = Path(__file__).parent / "seeds" / "ons_exogenous_snapshot.csv"
    if not snap_path.exists():
        return df
    snap = pd.read_csv(snap_path, index_col=0)
    snap.index = pd.PeriodIndex(snap.index, freq="Q")
    snap = snap.reindex(df.index)

    # Extrapolate each series beyond its last observation at its trailing
    # 4-quarter mean, NOT its last value. Many snapshot series are
    # non-seasonally-adjusted cash flows (e.g. CGT — capital gains tax, with
    # ~90% of receipts landing in Q1); holding the last value flat would freeze
    # a single seasonal quarter across the whole forecast (CGT enters at ~6x its
    # true annual rate if that quarter is the Q1 peak). The trailing annual
    # average removes the seasonal component while preserving the level, and is
    # harmless for smooth series and rates (mean ≈ last value).
    for col in snap.columns:
        valid = snap[col].dropna()
        if valid.empty:
            continue
        held = float(valid.iloc[-4:].mean())
        snap.loc[snap.index > valid.index[-1], col] = held

    # Internal gaps: forward-fill (interpolate a held assumption). Pre-sample:
    # a bounded back-fill only (limit=4) so a late-starting series is not
    # back-fabricated across decades of history it never observed.
    snap = snap.ffill().bfill(limit=4)
    add = {
        c: snap[c] for c in snap.columns if c not in df.columns or df[c].isna().all()
    }
    if add:
        df = pd.concat([df, pd.DataFrame(add, index=df.index)], axis=1)
        df = df.loc[:, ~df.columns.duplicated(keep="last")]
    return df


def ons_snapshot_edge(var: str):
    """Last genuine observation period of a vendored ONS snapshot series.

    The merged databank extrapolates every snapshot series beyond its last
    observation (trailing 4-quarter mean, see _merge_ons_snapshot), so the
    merged frame cannot distinguish data from held assumption. This reads the
    raw snapshot to report where the *published* observations end — callers
    that calibrate against observed data (e.g. the OSHH level anchor in
    full_solver) must not calibrate on the extrapolated tail. Returns a
    pd.Period, or None if the series is absent or the snapshot is missing.
    """
    snap_path = Path(__file__).parent / "seeds" / "ons_exogenous_snapshot.csv"
    if not snap_path.exists():
        return None
    snap = pd.read_csv(snap_path, index_col=0)
    if var not in snap.columns:
        return None
    valid = snap[var].dropna()
    if valid.empty:
        return None
    return pd.Period(valid.index[-1], freq="Q")


def _derive_variables(df: pd.DataFrame) -> pd.DataFrame:
    """Derive additional variables from identities."""
    # Real household disposable income
    if "HHDI" in df.columns and "PCE" in df.columns:
        df["RHHDI"] = df["HHDI"] / (df["PCE"] / 100)

    # Deflators from nominal/real ratios
    if "CONSPS" in df.columns and "CONS" in df.columns:
        df["PCONS"] = 100 * df["CONSPS"] / df["CONS"]

    if "IFPS" in df.columns and "IF" in df.columns:
        df["PIF"] = 100 * df["IFPS"] / df["IF"]

    if "XPS" in df.columns and "X" in df.columns:
        df["PX"] = 100 * df["XPS"] / df["X"]

    if "MPS" in df.columns and "M" in df.columns:
        df["PM"] = 100 * df["MPS"] / df["M"]

    if "CGGPS" in df.columns and "CGG" in df.columns:
        df["PCGG"] = 100 * df["CGGPS"] / df["CGG"]

    # Business investment deflator
    if "IBUS" in df.columns and "PIF" in df.columns:
        df["PIBUS"] = df["PIF"]  # Approximate

    # Investment share in business investment
    if "IBUS" in df.columns and "IF" in df.columns:
        df["IBUSSH"] = 100 * df["IBUS"] / df["IF"]

    # Employment in self-employment
    if "ETLFS" in df.columns and "EEES" in df.columns:
        df["ESLFS"] = df["ETLFS"] - df["EEES"]

    # Unemployment level: the model equations reference ULFS, while the EFO
    # table loads ULFSU (same LFS ILO unemployment level, thousands).
    if "ULFS" not in df.columns:
        if "ULFSU" in df.columns:
            df["ULFS"] = df["ULFSU"]
        elif "LFSUR" in df.columns and "ETLFS" in df.columns:
            df["ULFS"] = df["ETLFS"] * df["LFSUR"] / (100 - df["LFSUR"])

    # Current account as % of GDP
    if "CB" in df.columns and "GDPMPS" in df.columns:
        df["CBPCNT"] = 100 * df["CB"] / df["GDPMPS"]

    return df
