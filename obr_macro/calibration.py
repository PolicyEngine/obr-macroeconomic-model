"""Load OBR Economic and Fiscal Outlook data and calibrate model variables."""

from __future__ import annotations

import urllib.request
from pathlib import Path

import numpy as np
import openpyxl

from obr_macro.variables import Variables

# OBR November 2025 EFO download URLs
OBR_URLS = {
    "economy": "https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-economy/",
    "receipts": "https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-receipts/",
    "expenditure": "https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-expenditure/",
    "aggregates": "https://obr.uk/download/november-2025-economic-and-fiscal-outlook-detailed-forecast-tables-aggregates/",
}

OBR_FILENAMES = {
    "economy": "obr_efo_november_2025_economy.xlsx",
    "receipts": "obr_efo_november_2025_receipts.xlsx",
    "expenditure": "obr_efo_november_2025_expenditure.xlsx",
    "aggregates": "obr_efo_november_2025_aggregates.xlsx",
}


def ensure_obr_data(data_dir: str | Path | None = None) -> dict[str, Path]:
    """Download OBR EFO data files if not already present.

    Returns a dict mapping table type to file path.
    """
    if data_dir is None:
        data_dir = Path(__file__).parent.parent / "data"
    data_dir = Path(data_dir)
    data_dir.mkdir(parents=True, exist_ok=True)

    paths: dict[str, Path] = {}
    for key, filename in OBR_FILENAMES.items():
        path = data_dir / filename
        if not path.exists():
            url = OBR_URLS[key]
            print(f"Downloading OBR {key} tables from {url}...")
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req) as resp, open(path, "wb") as f:
                f.write(resp.read())
        paths[key] = path

    return paths


def _read_table(
    wb: openpyxl.Workbook,
    sheet: str,
    period_col: int,
    data_cols: dict[str, int],
    min_row: int = 5,
    max_row: int = 400,
) -> dict[str, dict[str, float]]:
    """Read quarterly data from an OBR table sheet.

    Returns {variable_name: {period_label: value}} for quarterly periods only.
    """
    ws = wb[sheet]
    result = {name: {} for name in data_cols}
    for r in range(min_row, max_row + 1):
        period = ws.cell(r, period_col).value
        if period is None:
            continue
        period = str(period)
        # Only keep quarterly data (e.g. '2024Q3')
        if len(period) != 6 or "Q" not in period:
            continue
        for name, col in data_cols.items():
            val = ws.cell(r, col).value
            if isinstance(val, (int, float)) and np.isfinite(val):
                result[name][period] = float(val)
    return result


def load_obr_efo(filepath: str | Path) -> dict[str, dict[str, float]]:
    """Load the OBR EFO economy supplementary tables.

    Returns a flat dict: {model_variable_name: {period_label: value}}.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    data: dict[str, dict[str, float]] = {}

    def merge(d: dict[str, dict[str, float]]) -> None:
        for k, v in d.items():
            data[k] = v

    # Table 1.7: Inflation indices (period in col B=2)
    # Growth rates in cols 3-12, index levels in cols 13-22
    merge(
        _read_table(
            wb,
            "1.7",
            period_col=2,
            data_cols={
                "RPI_yoy": 3,       # RPI yoy growth (%)
                "CPI_yoy": 5,       # CPI yoy growth (%)
                "PR": 13,           # RPI index (Jan 1987=100)
                "CPI": 15,          # CPI (2015=100)
                "CPIH": 16,         # CPIH (2015=100)
                "OOH": 17,          # OOH (2015=100)
                "PRMIP_raw": 18,    # Mortgage interest payments (Jan 1987=100)
                "_RENTS": 19,       # Actual rents for housing (2015=100)
                "PRENT": 20,        # Private rentals (2016=100)
                "PCE_raw": 21,      # Consumer expenditure deflator (2023=100)
                "PGDP_raw": 22,     # GDP deflator (2023=100)
            },
        )
    )

    # Table 1.9: Market-derived assumptions (period in col B=2, data from row 4)
    merge(
        _read_table(
            wb,
            "1.9",
            period_col=2,
            data_cols={
                "R": 3,         # Bank Rate (%)
                "RL": 4,        # Long-term gilt yield (%)
                "RMORT": 5,     # Average mortgage rate (%)
                "RDEP": 6,      # Deposit rate (%)
                "RX": 7,        # Sterling EER
                "RXD_GBP": 8,   # US$/£ exchange rate
                "PBRENT": 10,   # Oil prices ($)
                "FTSE": 12,     # Equity prices (FTSE All-Share)
            },
            min_row=4,
        )
    )

    # Table 1.1: Real GDP expenditure (period in col B=2)
    merge(
        _read_table(
            wb,
            "1.1",
            period_col=2,
            data_cols={
                "CONS": 3,      # Private consumption (real, £bn)
                "CGG": 4,       # Government consumption (real, £bn)
                "IF_real": 5,   # Fixed investment (real, £bn)
                "IBUS": 6,      # Business investment (real, £bn)
                "IH_real": 7,   # Private dwellings (real, £bn)
                "GGI_real": 8,  # General government investment (real, £bn)
                "X_real": 14,   # Exports (real, £bn)
                "M_real": 16,   # Imports (real, £bn)
                "GDPM_real": 18,  # Real GDP (£bn)
                "NOILGVA_real": 19,  # Non-oil GVA (real, £bn)
            },
        )
    )

    # Table 1.2: Nominal GDP expenditure (period in col B=2)
    merge(
        _read_table(
            wb,
            "1.2",
            period_col=2,
            data_cols={
                "CONSPS": 3,        # Private consumption (nominal, £bn)
                "CGGPS_nom": 4,     # Government consumption (nominal, £bn)
                "IFPS_nom": 5,      # Fixed investment (nominal, £bn)
                "XPS_nom": 11,      # Exports (nominal, £bn)
                "MPS_nom": 13,      # Imports (nominal, £bn)
                "GDPMPS_nom": 15,   # GDP at market prices (nominal, £bn)
                "GNIPS_nom": 16,    # Gross national income (nominal, £bn)
            },
        )
    )

    # Table 1.3: GDP income components (period in col B=2, data from row 4)
    merge(
        _read_table(
            wb,
            "1.3",
            period_col=2,
            data_cols={
                "LABOUR_INC": 3,    # Labour income (£bn)
                "FYCPR_nom": 4,     # Non-oil PNFC profits (£bn)
                "OTHER_INC": 5,     # Other income (£bn)
                "GVA_FC": 6,        # GVA at factor cost (£bn)
                "BPAPS_nom": 7,     # Taxes on products less subsidies (£bn)
                "GDPMPS_inc": 9,    # GDP (income measure, £bn)
            },
            min_row=4,
        )
    )

    # Table 1.6: Labour market (period in col B=2, data from row 4)
    merge(
        _read_table(
            wb,
            "1.6",
            period_col=2,
            data_cols={
                "ETLFS_m": 3,       # Employment (millions)
                "ER_pct": 4,        # Employment rate (%)
                "EMPLOYEES_m": 5,   # Employees (millions)
                "UNEMP_m": 6,       # ILO unemployment (millions)
                "LFSUR": 7,         # Unemployment rate (%)
                "PART16": 8,        # Participation rate (%)
                "AVH_raw": 9,       # Average hours worked
                "HWA_raw": 10,      # Total hours worked (millions)
                "COMP_EMP": 13,     # Compensation of employees (£bn)
                "WAGES_SAL": 14,    # Wages and salaries (£bn)
                "EMP_SOC": 15,      # Employers social contributions (£bn)
                "MI_raw": 16,       # Mixed income (£bn)
                "AWE_IDX": 18,      # AWE index (2008Q1=100)
            },
            min_row=4,
        )
    )

    # Table 1.12: Household disposable income (period in col B=2, data from row 4)
    merge(
        _read_table(
            wb,
            "1.12",
            period_col=2,
            data_cols={
                "HHDI_nom": 9,  # Household disposable income (£bn)
            },
            min_row=4,
        )
    )

    # Table 1.16: Housing market (period in col B=2, data from row 4)
    merge(
        _read_table(
            wb,
            "1.16",
            period_col=2,
            data_cols={
                "HPI": 3,           # House price index (Jan 2023=100)
                "HPI_yoy": 4,       # House price YoY change (%)
                "TRANS": 5,         # Residential property transactions (000s)
            },
            min_row=4,
        )
    )

    return data


def _set_var(
    v: Variables,
    name: str,
    series: dict[str, float],
    scale: float = 1.0,
    backfill: bool = True,
) -> None:
    """Set a model variable from an OBR quarterly series.

    For periods before OBR data starts, backfills with the earliest value.
    For periods after OBR data ends, forward-fills with the latest value.
    """
    if not series:
        return
    sorted_periods = sorted(series.keys())
    arr = v[name]

    # Fill known periods
    for period, val in series.items():
        try:
            t = v.period_to_idx(period)
            arr[t] = val * scale
        except (KeyError, IndexError):
            continue

    if backfill:
        earliest_val = series[sorted_periods[0]] * scale
        latest_val = series[sorted_periods[-1]] * scale

        # Find the model indices for earliest/latest OBR data
        try:
            t_first = v.period_to_idx(sorted_periods[0])
        except (KeyError, IndexError):
            t_first = 0
        try:
            t_last = v.period_to_idx(sorted_periods[-1])
        except (KeyError, IndexError):
            t_last = len(arr) - 1

        # Backfill before data
        arr[:t_first] = earliest_val
        # Forward-fill after data
        if t_last < len(arr) - 1:
            arr[t_last + 1 :] = latest_val


def calibrate_prices(v: Variables, data: dict[str, dict[str, float]]) -> None:
    """Set all price, interest rate and exchange rate variables from OBR data."""

    # CPI (2015=100) — model uses same basis
    _set_var(v, "CPI", data.get("CPI", {}))

    # CPIH (2015=100)
    _set_var(v, "CPIH", data.get("CPIH", {}))

    # OOH (2015=100)
    _set_var(v, "OOH", data.get("OOH", {}))

    # RPI index level (Jan 1987=100) — model stores as PR
    _set_var(v, "PR", data.get("PR", {}))

    # RPI year-on-year growth (%)
    _set_var(v, "RPI", data.get("RPI_yoy", {}))

    # Private rental prices (2016=100) — used for PRENT and CPIRENT
    _set_var(v, "PRENT", data.get("PRENT", {}))
    _set_var(v, "CPIRENT", data.get("PRENT", {}))

    # Consumer expenditure deflator (2023=100 in OBR)
    # Model uses this as an index — set directly
    _set_var(v, "PCE", data.get("PCE_raw", {}))

    # GDP deflator — OBR provides this as an index
    _set_var(v, "PGDP", data.get("PGDP_raw", {}))

    # Mortgage interest payments index
    _set_var(v, "PRMIP", data.get("PRMIP_raw", {}))

    # Interest rates (%) — direct mapping
    _set_var(v, "R", data.get("R", {}))
    _set_var(v, "RL", data.get("RL", {}))
    _set_var(v, "RMORT", data.get("RMORT", {}))
    _set_var(v, "RDEP", data.get("RDEP", {}))
    # Corporate bond rate: approximate as gilt yield + 1pp spread
    if "RL" in data:
        rocb_series = {p: val + 1.0 for p, val in data["RL"].items()}
        _set_var(v, "ROCB", rocb_series)

    # Exchange rates
    _set_var(v, "RX", data.get("RX", {}))
    # RXD in model is USD/GBP (how many dollars per pound)
    _set_var(v, "RXD", data.get("RXD_GBP", {}))

    # Oil prices ($/barrel)
    _set_var(v, "PBRENT", data.get("PBRENT", {}))


def calibrate_national_accounts(
    v: Variables, data: dict[str, dict[str, float]]
) -> None:
    """Set national accounts variables from OBR data.

    OBR reports in £bn; model uses £m. Scale = 1000.
    """
    S = 1000  # £bn → £m

    # Real GDP and components (Table 1.1, £bn constant prices)
    _set_var(v, "CONS", data.get("CONS", {}), scale=S)
    _set_var(v, "CGG", data.get("CGG", {}), scale=S)
    _set_var(v, "IF", data.get("IF_real", {}), scale=S)
    _set_var(v, "IBUS", data.get("IBUS", {}), scale=S)
    _set_var(v, "IH", data.get("IH_real", {}), scale=S)
    _set_var(v, "GGI", data.get("GGI_real", {}), scale=S)
    _set_var(v, "GGIX", data.get("GGI_real", {}), scale=S)
    _set_var(v, "GDPM", data.get("GDPM_real", {}), scale=S)
    _set_var(v, "TRGDP", data.get("GDPM_real", {}), scale=S)  # Trend = actual for now
    _set_var(v, "X", data.get("X_real", {}), scale=S)
    _set_var(v, "M", data.get("M_real", {}), scale=S)

    # Nominal GDP and components (Table 1.2, £bn current prices)
    _set_var(v, "CONSPS", data.get("CONSPS", {}), scale=S)
    _set_var(v, "CGGPS", data.get("CGGPS_nom", {}), scale=S)
    _set_var(v, "CGGPSPSF", data.get("CGGPS_nom", {}), scale=S)
    _set_var(v, "IFPS", data.get("IFPS_nom", {}), scale=S)
    _set_var(v, "GDPMPS", data.get("GDPMPS_nom", {}), scale=S)
    _set_var(v, "GNIPS", data.get("GNIPS_nom", {}), scale=S)
    _set_var(v, "XPS", data.get("XPS_nom", {}), scale=S)
    _set_var(v, "MPS", data.get("MPS_nom", {}), scale=S)

    # GDP income components (Table 1.3, £bn)
    _set_var(v, "FYCPR", data.get("FYCPR_nom", {}), scale=S)
    _set_var(v, "BPAPS", data.get("BPAPS_nom", {}), scale=S)

    # Derive FYEMP from labour income + employer social contributions
    if "COMP_EMP" in data:
        _set_var(v, "FYEMP", data["COMP_EMP"], scale=S)

    # Household disposable income (Table 1.12, £bn)
    _set_var(v, "HHDI", data.get("HHDI_nom", {}), scale=S)


def calibrate_labour_market(
    v: Variables, data: dict[str, dict[str, float]]
) -> None:
    """Set labour market variables from OBR data."""

    # Employment (OBR gives millions, model uses thousands)
    if "ETLFS_m" in data:
        etlfs = {p: val * 1000 for p, val in data["ETLFS_m"].items()}
        _set_var(v, "ETLFS", etlfs)
        _set_var(v, "ET", etlfs)

    if "EMPLOYEES_m" in data:
        employees = {p: val * 1000 for p, val in data["EMPLOYEES_m"].items()}
        # EMS = market sector employment ≈ employees - public sector
        # For now, approximate as total employees
        _set_var(v, "EMS", employees)

    # Unemployment rate (%)
    _set_var(v, "LFSUR", data.get("LFSUR", {}))

    # Participation rate (%)
    _set_var(v, "PART16", data.get("PART16", {}))

    # Employment rate (%)
    _set_var(v, "ER", data.get("ER_pct", {}))

    # Average hours worked
    if "AVH_raw" in data:
        _set_var(v, "AVH", data["AVH_raw"])

    # Total hours worked (OBR gives millions per week)
    if "HWA_raw" in data:
        _set_var(v, "HWA", data["HWA_raw"])

    # Mixed income (£bn → £m)
    _set_var(v, "MI", data.get("MI_raw", {}), scale=1000)

    # Employer social contributions (£bn → £m)
    _set_var(v, "EMPSC", data.get("EMP_SOC", {}), scale=1000)

    # Wages and salaries (£bn → £m) — this is WFP
    _set_var(v, "WFP", data.get("WAGES_SAL", {}), scale=1000)


def calibrate_housing(v: Variables, data: dict[str, dict[str, float]]) -> None:
    """Set housing market variables from OBR data."""
    # House price index (Jan 2023=100)
    _set_var(v, "APH", data.get("HPI", {}))


def _read_annual_fiscal(
    wb: openpyxl.Workbook,
    sheet: str,
    row_map: dict[str, int],
    label_col: int = 2,
    first_data_col: int = 4,
    year_row: int = 5,
    max_col: int = 11,
) -> dict[str, dict[str, float]]:
    """Read annual fiscal data and convert to quarterly series.

    OBR fiscal tables use financial years (e.g. '2025-26'). We spread each
    annual value evenly across the four quarters of the financial year:
    Q2, Q3, Q4 of the start year and Q1 of the end year.

    row_map: {model_variable_name: row_number}
    Returns {variable_name: {period_label: value}} with quarterly periods.
    """
    ws = wb[sheet]

    # Read the financial year headers from the year_row
    fy_cols: list[tuple[str, int]] = []
    for c in range(first_data_col, max_col + 1):
        fy = ws.cell(year_row, c).value
        if fy is None:
            continue
        fy = str(fy).strip()
        # Expect format like '2025-26'
        if len(fy) >= 5 and "-" in fy:
            fy_cols.append((fy, c))

    result: dict[str, dict[str, float]] = {}

    for var_name, row in row_map.items():
        series: dict[str, float] = {}
        for fy, col in fy_cols:
            val = ws.cell(row, col).value
            if not isinstance(val, (int, float)):
                continue
            if not np.isfinite(val):
                continue

            # Parse financial year: '2025-26' → start_year=2025
            try:
                start_year = int(fy[:4])
            except (ValueError, IndexError):
                continue

            # Quarterly value = annual / 4
            qval = float(val) / 4.0

            # Financial year runs Q2-Q4 of start_year, Q1 of start_year+1
            for q_label in [
                f"{start_year}Q2",
                f"{start_year}Q3",
                f"{start_year}Q4",
                f"{start_year + 1}Q1",
            ]:
                series[q_label] = qval

        result[var_name] = series

    return result


def load_obr_fiscal(
    receipts_path: str | Path,
    expenditure_path: str | Path,
    aggregates_path: str | Path,
) -> dict[str, dict[str, float]]:
    """Load OBR EFO fiscal supplementary tables (receipts, expenditure, aggregates).

    Returns a flat dict: {model_variable_name: {period_label: value}}.
    All values are in £bn (annual), converted to quarterly.
    """
    data: dict[str, dict[str, float]] = {}

    def merge(d: dict[str, dict[str, float]]) -> None:
        for k, v in d.items():
            data[k] = v

    # --- Receipts (Table 3.9: current receipts on cash basis) ---
    wb_r = openpyxl.load_workbook(str(receipts_path), data_only=True)
    merge(
        _read_annual_fiscal(
            wb_r,
            "3.9",
            row_map={
                "INCTAXG": 7,       # Income tax (gross of tax credits)
                "EMPNIC": 12,       # National insurance contributions
                "VREC": 13,         # VAT
                "CETAX_CT": 14,     # Corporation tax
                "PRT": 20,          # Petroleum revenue tax
                "TXFUEL": 21,       # Fuel duties
                "CGT": 22,          # Capital gains tax
                "INHT": 23,         # Inheritance tax
                "TSD": 24,          # Stamp duty land tax
                "TXSTSH": 26,       # Stamp taxes on shares
                "TXTOB": 27,        # Tobacco duties
                "TXALC_SP": 28,     # Spirits duties
                "TXALC_WN": 29,     # Wine duties
                "TXALC_BR": 30,     # Beer and cider duties
                "APD": 31,          # Air passenger duty
                "IPT": 32,          # Insurance premium tax
                "CCL": 33,          # Climate change levy
                "CUST": 37,         # Customs duties
                "BLEVY": 38,        # Bank levy
                "VED": 52,          # Vehicle excise duties
                "NNDRA": 53,        # Business rates
                "CC": 54,           # Council tax
                "PSCR": 64,         # Total current receipts
            },
            first_data_col=4,
            year_row=5,
            max_col=10,
        )
    )

    # Table 3.4: Income tax and NICs detailed breakdown
    merge(
        _read_annual_fiscal(
            wb_r,
            "3.4",
            row_map={
                "TYEM_PAYE": 8,     # PAYE
                "TYEM_SA": 9,       # Self assessment
                "NIC_EE": 17,       # Class 1 employee NICs
                "NIC_ER": 18,       # Class 1 employer NICs
            },
            first_data_col=3,
            year_row=5,
            max_col=9,
        )
    )

    # --- Expenditure ---
    wb_e = openpyxl.load_workbook(str(expenditure_path), data_only=True)

    # Table 4.9: Welfare spending
    merge(
        _read_annual_fiscal(
            wb_e,
            "4.9",
            row_map={
                "WELFARE_CAP": 36,      # Total welfare cap spending
                "STATE_PENSION": 40,    # State pension
                "WELFARE_TOTAL": 47,    # Total welfare spending
            },
            first_data_col=4,
            year_row=5,
            max_col=10,
        )
    )

    # --- Aggregates ---
    wb_a = openpyxl.load_workbook(str(aggregates_path), data_only=True)

    # Table 6.3: General government transactions
    merge(
        _read_annual_fiscal(
            wb_a,
            "6.3",
            row_map={
                "GG_INC_WEALTH": 7,     # Taxes on income and wealth
                "GG_PROD_IMP": 8,       # Taxes on production and imports
                "GG_OTHER_TAX": 9,      # Other current taxes
                "GG_CAP_TAX": 10,       # Taxes on capital
                "GG_SOC_CONT": 11,      # Compulsory social contributions
                "GG_GOS": 12,           # Gross operating surplus
                "GG_RENT_TR": 13,       # Rent and other current transfers
                "GG_INT_DIV_PRIV": 14,  # Interest & dividends from private
                "GG_INT_DIV_PUB": 15,   # Interest & dividends from public
                "GG_CURR_RECEIPTS": 16, # Total current receipts
                "GG_CONSUMPTION": 19,   # Consumption expenditure
                "GG_SUBSIDIES": 20,     # Subsidies
                "GG_NET_SOC_BEN": 21,   # Net social benefits
                "GG_CURR_GRANTS_ABR": 22, # Net current grants abroad
                "GG_OTHER_GRANTS": 24,  # Other current grants
                "GG_INT_PAID": 26,      # Interest and dividends paid
                "GG_CURR_EXP": 27,      # Total current expenditure
                "GG_DEPRECIATION": 28,  # Depreciation
                "GG_GDFCF": 32,         # Gross domestic fixed capital formation
                "GG_NET_BORROW": 39,    # Net borrowing
                "CG_NET_BORROW": 41,    # Central govt net borrowing
                "LA_NET_BORROW": 42,    # Local authority net borrowing
            },
            first_data_col=3,
            year_row=5,
            max_col=9,
        )
    )

    # Table 6.1: Expenditure by sector
    merge(
        _read_annual_fiscal(
            wb_a,
            "6.1",
            row_map={
                "CG_CONSUMPTION": 10,   # CG consumption
                "CG_SUBSIDIES": 11,     # CG subsidies
                "CG_SOC_BEN": 12,       # CG net social benefits
                "CG_GRANTS_ABR": 13,    # CG net current grants abroad
                "CG_INT_GRANTS": 14,    # CG current grants within public sector
                "CG_OTHER_GRANTS": 15,  # CG other current grants
                "CG_INT_PAID": 17,      # CG interest and dividends paid
                "CG_GDFCF": 21,         # CG GDFCF
                "CG_CAP_GRANTS": 25,    # CG capital grants to private sector
                "LA_CONSUMPTION": 32,   # LA consumption
                "LA_SUBSIDIES": 33,     # LA subsidies
                "LA_SOC_BEN": 34,       # LA net social benefits
                "LA_INT_PAID": 38,      # LA interest and dividends paid
                "LA_GDFCF": 42,         # LA GDFCF
                "LA_CAP_GRANTS": 46,    # LA capital grants to private sector
                "PSCE": 96,             # Public sector current expenditure
                "PSNI": 97,             # Public sector net investment
            },
            first_data_col=4,
            year_row=5,
            max_col=10,
        )
    )

    return data


def calibrate_fiscal(
    v: Variables, data: dict[str, dict[str, float]]
) -> None:
    """Set fiscal variables from OBR fiscal tables.

    OBR fiscal data is in £bn annual, already converted to quarterly £bn
    by load_obr_fiscal. Model uses £m, so scale = 1000.
    """
    S = 1000  # £bn → £m

    # Tax receipts
    # Income tax: TYEM in model = total income tax on households
    # Use PAYE as main proxy for TYEM (employed income tax)
    _set_var(v, "TYEM", data.get("TYEM_PAYE", {}), scale=S)

    # Self-assessment goes into TSEOP (self-employment tax)
    _set_var(v, "TSEOP", data.get("TYEM_SA", {}), scale=S)

    # Total income tax (gross) for reference
    _set_var(v, "INCTAXG", data.get("INCTAXG", {}), scale=S)

    # NICs
    if "NIC_ER" in data:
        _set_var(v, "EENIC", data["NIC_ER"], scale=S)
    if "NIC_EE" in data:
        _set_var(v, "EMPNIC", data["NIC_EE"], scale=S)

    # VAT
    _set_var(v, "VREC", data.get("VREC", {}), scale=S)

    # Corporation tax
    _set_var(v, "CETAX", data.get("CETAX_CT", {}), scale=S)

    # Fuel duties
    _set_var(v, "TXFUEL", data.get("TXFUEL", {}), scale=S)

    # Tobacco duties
    _set_var(v, "TXTOB", data.get("TXTOB", {}), scale=S)

    # Alcohol duties (combine spirits + wine + beer)
    if "TXALC_SP" in data and "TXALC_WN" in data and "TXALC_BR" in data:
        combined = {}
        for p in data["TXALC_SP"]:
            val = data["TXALC_SP"].get(p, 0) + data["TXALC_WN"].get(p, 0) + data["TXALC_BR"].get(p, 0)
            combined[p] = val
        _set_var(v, "TXALC", combined, scale=S)

    # Capital gains tax
    _set_var(v, "CGT", data.get("CGT", {}), scale=S)

    # Inheritance tax
    _set_var(v, "INHT", data.get("INHT", {}), scale=S)

    # Stamp duty land tax
    _set_var(v, "TSD", data.get("TSD", {}), scale=S)

    # Council tax
    _set_var(v, "CC", data.get("CC", {}), scale=S)

    # Business rates
    _set_var(v, "NNDRA", data.get("NNDRA", {}), scale=S)

    # Customs duties
    _set_var(v, "CUST", data.get("CUST", {}), scale=S)

    # Climate change levy
    _set_var(v, "CCL", data.get("CCL", {}), scale=S)

    # Bank levy
    _set_var(v, "BLEVY", data.get("BLEVY", {}), scale=S)

    # Vehicle excise duties
    _set_var(v, "VED", data.get("VED", {}), scale=S)

    # Petroleum revenue tax
    _set_var(v, "PRT", data.get("PRT", {}), scale=S)

    # Total current receipts
    _set_var(v, "PSCR", data.get("PSCR", {}), scale=S)

    # Social benefits
    # CGSB = central government social benefits (DWP social security)
    _set_var(v, "CGSB", data.get("CG_SOC_BEN", {}), scale=S)

    # LASBHH = local authority social benefits to households
    _set_var(v, "LASBHH", data.get("LA_SOC_BEN", {}), scale=S)

    # Total welfare spending for reference
    _set_var(v, "PUBSTIW", data.get("WELFARE_TOTAL", {}), scale=S)

    # Government consumption
    _set_var(v, "CGGPSPSF", data.get("GG_CONSUMPTION", {}), scale=S)

    # Government investment (GDFCF)
    _set_var(v, "CGIPS", data.get("CG_GDFCF", {}), scale=S)
    _set_var(v, "LAIPS", data.get("LA_GDFCF", {}), scale=S)

    # Government subsidies
    _set_var(v, "CGSUBP", data.get("CG_SUBSIDIES", {}), scale=S)
    _set_var(v, "LASUBP", data.get("LA_SUBSIDIES", {}), scale=S)

    # Government interest paid (debt interest)
    _set_var(v, "PSINTR", data.get("GG_INT_PAID", {}), scale=S)

    # Capital grants to private sector
    _set_var(v, "CGKTA", data.get("CG_CAP_GRANTS", {}), scale=S)

    # Other grants
    _set_var(v, "CGOTR", data.get("CG_OTHER_GRANTS", {}), scale=S)

    # Net borrowing
    _set_var(v, "PSNBCY", data.get("GG_NET_BORROW", {}), scale=S)

    # Compulsory social contributions (general government)
    _set_var(v, "EMPSC", data.get("GG_SOC_CONT", {}), scale=S)

    # Depreciation
    _set_var(v, "GGFCD", data.get("GG_DEPRECIATION", {}), scale=S)


def calibrate_all(v: Variables, data: dict[str, dict[str, float]]) -> None:
    """Apply all calibration functions."""
    calibrate_prices(v, data)
    calibrate_national_accounts(v, data)
    calibrate_labour_market(v, data)
    calibrate_housing(v, data)
